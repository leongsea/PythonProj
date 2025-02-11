# 这是一个示例 Python 脚本。

# 按 Shift+F10 执行或将其替换为您的代码。
# 按 Double Shift 在所有地方搜索类、文件、工具窗口、操作和设置。
import requests
import openpyxl
import json
from datetime import datetime


class userDto:
    def __init__(self, Mac, IP, CPU, UserName, Department, Location, Area, EquipmentType):
        self.id = 0
        self.Mac = Mac
        self.IP = IP
        self.CPU = CPU
        self.UserName = UserName
        self.Department = Department
        self.Location = Location
        self.Area = Area
        self.EquipmentType = EquipmentType
        self.InventoryIn = "2024-12-31"
        self.UpdateTime = None

    def userDto_Serializer(self):
        return {
            "id": 0,
            "userName": self.UserName,
            "ip": self.IP,
            "mac": self.Mac,
            "cpu": self.CPU,
            "equipmentType": self.EquipmentType,
            "area": self.Area,
            "location": self.Location,
            "department": self.Department,
            "inventoryIn": self.InventoryIn,
            "updateTime": self.UpdateTime
        }


def print_hi():
    wb = openpyxl.load_workbook(r"C:\Users\Leong\Desktop\新建 Microsoft Excel 工作表.xlsx")
    ws = wb['Sheet1']
    userDict = []
    for i in range(2, 131):
        user = userDto(Mac=ws[f'B{i}'].value,
                       IP=ws[f'C{i}'].value,
                       CPU=ws[f'D{i}'].value,
                       UserName=ws[f'E{i}'].value,
                       Department=ws[f'F{i}'].value,
                       Location=ws[f'G{i}'].value,
                       Area=ws[f'H{i}'].value,
                       EquipmentType=ws[f'I{i}'].value
                       )
        userDict.append(user.userDto_Serializer())
    wb.close()
    return userDict


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    userdtoDict = print_hi()
    global userdto
    headers = {'Content-Type': 'application/json',
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36'}
    for userdto in userdtoDict:
        json_user = json.dumps(userdto)
        try:
            res = requests.post("http://localhost:5157/api/OfficeEquipments", data=json_user, headers=headers)
            print(f"{userdto['ip']}--->{res.text}")
        except Exception as e:
            print(f"{userdto['ip']} 存入失败!")

# 访问 https://www.jetbrains.com/help/pycharm/ 获取 PyCharm 帮助
